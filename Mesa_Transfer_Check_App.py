from openpyxl import load_workbook
import tkinter as tk
from tkinter import font


######################EXCEL WORK############################################
mesa = load_workbook('Mesa Transfer Items.xlsx')

hojas = []
for sheet in mesa.sheetnames:
    hojas.append(mesa[sheet])

b1 = hojas[0]
b2 = hojas[1]
b3 = hojas[2]
b4 = hojas[3]
b5 = hojas[4]
not_in = hojas[5]
in_lot = hojas[6]


######BIN 1 Items with Quantity
b1_items = []
for i in b1['A']:
    b1_items.append(str(i.value))

b1_qty = []
for i in b1['B']:
    b1_qty.append(i.value)

b1_item_qty = [(b1_items[i],b1_qty[i]) for i in range(0, len(b1_items))]


######BIN 2 Items with Quantity
b2_items = []
for i in b2['A']:
    b2_items.append(str(i.value))

b2_qty = []
for i in b2['B']:
    b2_qty.append(i.value)

b2_item_qty = [(b2_items[i],b2_qty[i]) for i in range(0, len(b2_items))]


######BIN 3 Items with Quantity
b3_items = []
for i in b3['A']:
    b3_items.append(str(i.value))

b3_qty = []
for i in b3['B']:
    b3_qty.append(i.value)

b3_item_qty = [(b3_items[i],b3_qty[i]) for i in range(0, len(b3_items))]


######BIN 4 Items with Quantity
b4_items = []
for i in b4['A']:
    b4_items.append(str(i.value))

b4_qty = []
for i in b4['B']:
    b4_qty.append(i.value)

b4_item_qty = [(b4_items[i],b4_qty[i]) for i in range(0, len(b4_items))]


######BIN 5 Items with Quantity
b5_items = []
for i in b5['A']:
    b5_items.append(str(i.value))

b5_qty = []
for i in b5['B']:
    b5_qty.append(i.value)

b5_item_qty = [(b5_items[i],b5_qty[i]) for i in range(0, len(b5_items))]


###### Lotus Notes Mesa List and Missing Items in Lotus Notes Mesa List
not_in_l = [(str(i.value)) for i in not_in['A']]
in_lot_m = [(str(i.value)) for i in in_lot['A']]


########################FUNCTIONS######################################
def format_response(entry,f, qty):
    w_bin = f
    v = "ITEM DID NOT MAKE INITIAL DATA TRANSFER \nNOT AVAILABLE IN LOTUS NOTES MESA TRANSFER DATABASE \nMUST CONTACT I.T. FOR FURTHER SUPPORT"
    qty_r = str(qty)
    final_str = 'KB%s \n\nLOCATED IN %s - QTY: %s \n\n%s ' % (entry,w_bin, qty_r,v)
    return final_str

def format_response2(entry,f,qty):
    c_bin = f
    w = "AVAILABLE IN LOTUS NOTES MESA TRANSFER DATABASE"
    qty_s = str(qty)
    last_str = 'KB%s \n\nLOCATED IN %s - QTY: %s \n\n%s' % (entry,c_bin, qty_s,w)
    return last_str


def kb_search(entry):
    
    if entry in b1_items:
        x = b1_items.index(entry)
        qty = b1_qty[x]
        f = b1.title 

        if entry in not_in_l:                
            label['text'] = format_response(entry,f,qty)     
        elif entry in in_lot_m:
            label['text'] = format_response2(entry,f,qty)

    elif entry in b2_items:
        x = b2_items.index(entry)
        qty = b2_qty[x]
        f = b2.title

        if entry in not_in_l:
            label['text'] = format_response(entry,f,qty)
        elif entry in in_lot_m:         
            label['text'] = format_response2(entry,f,qty)

    elif entry in b3_items:
        x = b3_items.index(entry)
        qty = b3_qty[x]
        f = b3.title

        if entry in not_in_l:
           label['text'] = format_response(entry,f,qty)
        elif entry in in_lot_m:
            label['text'] = format_response2(entry,f,qty)

    elif entry in b4_items:
        x = b4_items.index(entry)
        qty = b4_qty[x]
        f = b4.title

        if entry in not_in_l:            
            label['text'] = format_response(entry,f, qty)
        elif entry in in_lot_m:
            label['text'] = format_response2(entry,f, qty)

    elif entry in b5_items:
        x = b5_items.index(entry)
        qty = b5_qty[x]
        f = b5.title
        
        if entry in not_in_l:          
            label['text'] = format_response(entry,f, qty)
        elif entry in in_lot_m:
            label['text'] = format_response2(entry,f, qty)

    elif entry in in_lot_m:
        label['text'] = "KB"+ entry + "\n\nITEM IS AVAILABLE IN LOTUS NOTES MESA TRANSFER DATABASE, BUT WAS NOT TRANSFERRED TO BENSENVILLE"

    else:
        label['text'] = "KB"+ entry + "\n\nITEM DOES NOT EXIST IN MESA TRANSFER LIST AND DOES NOT EXIST IN LOTUS NOTES MESA TRANSFER DATABSE"


#####################GUI WINDOW#############################################
H = 700
W = 925


root =tk.Tk()
root.title("Mesa Transfer Items Search")

canvas = tk.Canvas(root, bg = "#808080" , height = H, width = W)
canvas.pack(fill = "both", expand=True)


frame = tk.Frame(root, bg='#b30000', bd=10)
frame.place(relx=0.5, rely=0.15, relwidth = 0.6, relheight = 0.09, anchor='n')

entry = tk.Entry(frame, bg ="#d9d9d9",font=(None, 18), relief = "sunken")
entry.place(relwidth=0.65, relheight=1)

button = tk.Button(frame, text="Click to Search", font = 40,relief = "raised",command=lambda:kb_search(entry.get()))
button.place(relx=0.7, relheight=1, relwidth=0.3)

mid_label = tk.Label(root, bg ="#808080" ,text= "Enter a Single KB Number. Example: 471, 8651FRT, 8651BK, 8295INSERT", font = ('Calibri Light',10,'italic', 'bold') )
mid_label.place(relx =0.45, rely = 0.25, anchor = 'n') 


lower_frame = tk.Frame(root, bg='#b30000', bd=10)
lower_frame.place(relx=0.5, rely=0.35, relwidth=0.9, relheight=0.4, anchor='n')


label = tk.Label(lower_frame,bg ="#d9d9d9"  ,font=('Calibri', 12, 'bold'))
label.place(relwidth=1, relheight=1)


bottom_label = tk.Label(root, bg = "#808080" ,text = "Owned and Developed by Gerardo Torres \nJTCreations", font = ('Calibri Light',8))
bottom_label.place(relx = 0.5, rely = 0.98, anchor ='s')

root.mainloop()

